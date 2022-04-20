# -*- coding: utf-8 -*-
# PyTabsam
# @author: sszsth, sszgrm

import json
import pandas as pd
import os # operating system
import re # regular expressions
import openpyxl
from openpyxl.styles import Font
import shutil
import datetime
from copy import copy

# Leere Listen vorbereiten
data_coll  = pd.DataFrame([],dtype=pd.StringDtype())
count_sheet = 0
data_sheet = pd.DataFrame([],dtype=pd.StringDtype())
count_expl = 0
data_expl  = pd.DataFrame([],dtype=pd.StringDtype())
count_foot = 0
data_foot  = pd.DataFrame([],dtype=pd.StringDtype())
# Global variable from configuration
path_input = ""
path_output = ""

# Function tolog
# Write logging information
def tolog(level, text):
  dateTimeObj = datetime.datetime.now()
  timestamp = dateTimeObj.strftime("%Y-%m-%d %H:%M:%S%z")
  print(level + " (" + timestamp + "): " + text)


# Function read_config
# Read the configuration file
def read_config():
  with open('config.json', 'r', encoding="utf-8") as f:
    config = json.load(f)
    list_coll = []
    global data_coll, path_input, path_output
    
    path_input  = config['path_input']
    path_output = config['path_output']
    
    for key in config:
      conf_value = config[key]
      if key == "collection":
        # The collections are provided as a list. Read them an add them do the Dataframe data_coll
        for i in range(len(conf_value)):
          coll_elem = conf_value[i]
          pk = i+1
          input_fullpath = path_input + "/" + coll_elem["input_subpath"]
          elem_list_coll = [pk, coll_elem["title"], input_fullpath, coll_elem["output_filename"]]
          list_coll.append(elem_list_coll)
          data_coll = pd.DataFrame(list_coll, columns = ['id', 'title' , 'input_path', 'output_filename'])


# Function read_coll_dir
# Open each collection path and scan for all files in it. Match names to expl. and sheets.
def read_coll_dir():
  global data_coll, count_sheet, data_sheet, count_expl, data_expl
  list_sheet = []
  count_sheet = 0
  list_expl = []
  count_expl = 0
  data_coll = data_coll.reset_index()  # make sure indexes pair with number of rows
  for index, row in data_coll.iterrows():
    collection_id   = row['id']
    collection_path = row['input_path']
    tolog("INFO", "Reading collection from: " + collection_path)


    # Get list of all files only in the given directory
    list_of_files = filter( lambda x: os.path.isfile(os.path.join(collection_path, x)),
                            os.listdir(collection_path) )
    # Create a list of files in directory along with the size
    files_with_size = [ (file_name, os.stat(os.path.join(collection_path, file_name)).st_size) 
                        for file_name in list_of_files  ]
    # Iterate over list of files along with size 
    # and print them one by one.
    for filename, size in files_with_size:
      if size>600000:
        tolog("WARNING", "File " + filename + " in " + collection_path + " is larger than 600KB and may be damaged. It will be ignored.")
      else:
        is_expl = re.match(r'T_\d+.\d+.\d+.Erläuterungen.xlsm', filename)
        # Regular Expressen for an excel file containing an explanation
        # Example: T_02.02.0.Erläuterungen.xlsm
        is_sheet = re.match(r'(T_|T_G)\d+.\d+.\d+(.xlsm|.\d+.xlsm|\w.xlsm)', filename)
        # Regular Expressen for an excel file containing data
        # Examples: T_02.02.03.xlsm, T_02.02.01.2017.xlsm, T_G03.03.01.xlsm, T_G07.03.05a
  
        # The filename matches an explanation
        if is_expl:
          count_expl += 1
          elem_list_expl = [count_expl, collection_id, filename, collection_path]
          list_expl.append(elem_list_expl)
        # The filename matches a worksheet with data
        elif is_sheet:
          count_sheet += 1
          sheet_name = filename.replace('.xlsm', '').replace('_', '')
          elem_list_sheet = [count_sheet, collection_id, filename, collection_path, sheet_name]
          list_sheet.append(elem_list_sheet)
        else:
          tolog("WARNING", "File " + filename + " in " + collection_path + " has an invalid filename. It will be ignored.")
        
    data_expl  = pd.DataFrame(list_expl, columns = ['ID', 'FK_collection', 'filename', 'directory'])
    data_sheet = pd.DataFrame(list_sheet, columns = ['ID', 'FK_collection', 'filename', 'directory', 'sheet_name'])


# Function read_xls_metadata
# Read metadata of excel 
def read_xls_metadata(sheet_id):
  global data_sheet
  xls_filename = data_sheet.loc[data_sheet['ID'] == sheet_id, 'filename'].values[0]
  xls_directory = data_sheet.loc[data_sheet['ID'] == sheet_id, 'directory'].values[0]
  xls_path = xls_directory + "/" + xls_filename
  #print("Excel: "+ xls_path)
  wb = openpyxl.load_workbook(filename = xls_path)
  sheet_md = wb['Metadaten']
  colkey = ""
  for (row, col), source_cell in sheet_md._cells.items():
    #print(str(row) + " - "+ str(col) + ": " + str(source_cell._value))
    if col==1:
      colkey = str(source_cell._value).lower()
    # add values from the metadata sheet directly to the dataframe by ID
    # wfn stands for "without footnote" an can late be used in the table of contents
    # re.sub removes " #1" and "#1" (without trailing blanks) from strings
    if col==2:
      if colkey=="tabellencode":
        #print("Tabellencode: " + str(source_cell._value))
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['code']] = str(source_cell._value)
      if colkey=="tabellentitel":
        #print("Tabellentitel: " + str(source_cell._value))
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['title']] = str(source_cell._value)
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['title_wfn']] = re.sub('( #[0-9]+)|(#[0-9]+)', '', str(source_cell._value))
      if colkey=="tabellenuntertitel1":
        #print("Tabellenuntertitel1: " + str(source_cell._value))
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['subtitle1']] = str(source_cell._value)
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['subtitle1_wfn']] = re.sub('( #[0-9]+)|(#[0-9]+)', '', str(source_cell._value))
      if colkey=="tabellenuntertitel2":
        #print("Tabellenuntertitel2: " + str(source_cell._value))
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['subtitle2']] = str(source_cell._value)
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['subtitle2_wfn']] = re.sub('( #[0-9]+)|(#[0-9]+)', '', str(source_cell._value))
      if colkey=="quelle":
        #print("Quelle: " + str(source_cell._value))
        data_sheet.loc[data_sheet['ID'] == sheet_id, ['source']] = str(source_cell._value)


# Function read_xls_footnote
# Read footnote of excel 
def read_xls_footnote(sheet_id, list_foot):
  global data_sheet, count_foot, data_foot
  xls_filename = data_sheet.loc[data_sheet['ID'] == sheet_id, 'filename'].values[0]
  xls_directory = data_sheet.loc[data_sheet['ID'] == sheet_id, 'directory'].values[0]
  xls_path = xls_directory + "/" + xls_filename
  #print("Excel: "+ xls_path)
  wb = openpyxl.load_workbook(filename = xls_path)
  sheet_fn = wb['Fussnoten']
  coltype = ""
  colfnid = ""
  for (row, col), source_cell in sheet_fn._cells.items():
    #print(str(row) + " - "+ str(col) + ": " + str(source_cell._value))
    if col==1:
      coltype = str(source_cell._value).lower()
    if col==2:
      colfnid = str(source_cell._value).lower().strip()
    if col==3:
      if coltype=="<web>":
        count_foot += 1 
        #print("Fussnote Nr: " + colfnid + " = " + str(source_cell._value))
        elem_list_foot = [count_foot, sheet_id, colfnid, str(source_cell._value), "false"]
        list_foot.append(elem_list_foot)


# Function read_all_md_fn
# Read metadata and footnote sheets of all excel documents in sheet dataframe 
# As it is very inefficient to grow a dataframe, the footnotes are collectet
# in a list, which is then converted to a dataframe
def read_all_md_fn():
  global count_sheet, data_sheet, data_foot
  #print(count_sheet)
  sheet_id = 0
  list_foot = []
  while sheet_id < count_sheet:
    sheet_id += 1
    read_xls_metadata(sheet_id)
    read_xls_footnote(sheet_id, list_foot)
    data_foot = pd.DataFrame(list_foot, columns = ['ID', 'FK_sheet', 'code', 'text', 'used'])


# Function to_superscript
# Convert footnote codes to their equivalent in unicode superscript characters
# This function is limited to 25 footnotes
def to_superscript(matchobj):
    m =  matchobj.group(0)
    map = {'#1': '\u00B9', '#2': '\u00B2', '#3': '\u00B3', '#4': '\u2074', '#5': '\u2075',
           '#6': '\u2076', '#7': '\u2077', '#8': '\u2078', '#9': '\u2079', '#10': '\u00B9\u2070',
           '#11': '\u00B9\u00B9', '#12': '\u00B9\u00B2', '#13': '\u00B9\u00B3', 
           '#14': '\u00B9\u2074', '#15': '\u00B9\u2075',  '#16': '\u00B9\u2076',
           '#17': '\u00B9\u2077', '#18': '\u00B9\u2078',  '#19': '\u00B9\u2079', 
           '#20': '\u00B2\u2070', '#21': '\u00B2\u00B9', '#22': '\u00B2\u00B2',
           '#23': '\u00B2\u00B3', '#24': '\u00B2\u2074', '#25': '\u00B2\u2075'}
    return map[m]


# Function convert_footnote 
# Check if string contains reference to a footnote and if a related description exists
# Mark used footnotes in data_foot for later output
# Convert "#1" notation of input to nice superscript for output
def convert_footnote(sheet_id, input_string):
  global data_foot, data_sheet
  output_string = input_string
  # check if a sheet_id is provided. Explanations dont have footnotes and will provide a zero
  if sheet_id==0:
    return(input_string)
  # check if there is a footnote code in the input string (with or without leading space)
  elif re.search("( #[0-9]+)|(#[0-9]+)", input_string):
    # At least one footnote was found
    fn_refs = re.findall(" #[0-9]+|#[0-9]+", input_string)
    for fn_ref in fn_refs:
      # extract the number (if it has for example a leading space)
      match = re.search("(#[0-9]+)", fn_ref)
      fn_code = match.group(1);
      # check if footnote code exists in data_foot
      rslt_df = data_foot[(data_foot['FK_sheet'] == sheet_id) & (data_foot['code'] == fn_code)]
      if rslt_df.empty:
        sheetname = data_sheet.loc[(data_sheet['ID'] == sheet_id), "filename"].iloc[0]
        tolog("WARNING", "The footnote " + fn_code + " was referenced, but not found in the worksheet Fussnoten of " + sheetname)
      else:
        # set "used" to true, to check later (when creating the table) if all the footnotes were used 
        data_foot.loc[(data_foot['FK_sheet'] == sheet_id) & (data_foot['code'] == fn_code), "used"] = "true"
        # replace the hash-sign with superscript code for nicer Excel ouput
        pattern = "#(" + fn_code[1:] + ")"
        output_string = re.sub(pattern, to_superscript, output_string)
    return(output_string)
  else:
    # No reference to a footnote found
    return(input_string)


# Function prepare_footnotes 
# Check if all footnotes from the metadata have been referenced 
# Returns a list with the footnote items for the output worksheet
def prepare_footnotes(sheet_id):
  list_fn = []
  for index, row in data_foot[(data_foot['FK_sheet'] == sheet_id)].iterrows():
    if row['used']=="false":
      sheetname = data_sheet.loc[(data_sheet['ID'] == sheet_id), "filename"].iloc[0]
      tolog("WARNING", "The footnote " +  row['code'] + " is not used, but was found in the worksheet Fussnoten of " + sheetname)
    else:
      code_sup = re.sub("(#[0-9]+)", to_superscript, row['code'])
      list_fn.append(code_sup + " " + row['text'])
  return(list_fn)


# Function write_footnotes
# Add footnotes to the end of the worksheet
def write_footnotes(dest_ws, list_wsfn, row_start):
  if len(list_wsfn)==0:
    return(row_start)
  else:
    for i in range(len(list_wsfn)):
      dest_ws.cell(row=row_start+i, column=1).value = list_wsfn[i]
      dest_ws.cell(row=row_start+i, column=1).font = Font(name='Arial', size=8)
  return(row_start+i)


# Function create_tabsam 
# Create and generate the destination excel files
# If the destination files already exists, they will be overwritten
def create_tabsam():
  global path_output
  coll_ID = 0
  for index, row in data_coll.iterrows():
    # Copy the template into the existing xlsx files according to the collection
    coll_ID = row['id']
    tolog("INFO", "Writing collection to: " + row['output_filename'])
    filename = path_output + '/' + row['output_filename']
    shutil.copy('VorlageTabsam.xlsx', filename)
    
    # Open Excel file
    file = openpyxl.load_workbook(filename)

    # Set metadata property title of file
    file.properties.title = row['title']

    # Fill in the creation date into the worksheet 'Inhalt' 
    sheet = file["Inhalt"]
    today = datetime.date.today()
    create_date = "Erstellt am: " + today.strftime("%d.%m.%Y")
    sheet.cell(row=2, column=3).value = create_date
    file.save(filename)
      
    # Create and fill in the worksheet 'Erläuterungen' if exists
    create_worksheet_expl(coll_ID, filename)

    # Create and fill all worksheet according to the dataframe sheet
    create_worksheets(coll_ID, filename)
  

# Function create_worksheet_expl
# Read the data from the source worksheet "Internet" and write it to the destination worksheet
# Copy all the format of the source worksheet "Internet"
# Write the titel in the table of content in the worksheet "Inhalt"
# Save the xlsx file
def create_worksheet_expl(coll_ID, dest_file):
  for index, row in data_expl.iterrows():
    if coll_ID == row['FK_collection']:
      # Opening the source xlsx
      source_xlsx = row['directory'] + "/" + row['filename']
      source_wb = openpyxl.load_workbook(source_xlsx)
      source_ws = source_wb["Internet"]
      
      # Opening the destination xlsx and create the new worksheet
      dest_wb = openpyxl.load_workbook(dest_file)
      dest_ws = dest_wb.create_sheet("Erläuterungen")

      # Read the data from the source worksheet "Internet" and write it to the destination worksheet
      # Copy all the format of the source worksheet "Internet"
      last_row = read_write_data(source_ws, dest_ws, 0, 0)

      # Write "Erläuterungen" in the table of contents
      content = dest_wb["Inhalt"]
      content.cell(row=10, column=1).value = "Erläuterungen"
      content.cell(row=10, column=2).value = ""
  
      # Saving the destination xlsx file
      dest_wb.save(dest_file)
      

# Function create_worksheets
# Write the header of the sheet
# Read the data from the source worksheet "Internet" and write it to the destination worksheet
# Copy all the format of the source worksheet "Internet"
# Set the uniform row height to 12.75 for the common worksheet
# Write the titel in the table of content in the worksheet "Inhalt"
# Save the xlsx file
def create_worksheets(coll_ID, dest_file):
  # Set the uniform row height in the destination worksheet
  row_height = 11.25
  
  # Set the counter for the table of content
  wb = openpyxl.load_workbook(dest_file, read_only=True)
  if 'Erläuterungen' in wb.sheetnames:
    count_toc = 11
  else:
    count_toc = 10
        
  for index, row in data_sheet.iterrows():
    if coll_ID == row['FK_collection']:

      # Opening the source xlsx
      source_xlsx = row['directory'] + "/" + row['filename']
      source_wb = openpyxl.load_workbook(source_xlsx)
      source_ws = source_wb["Internet"]

      # Opening the destination xlsx and create the new worksheet
      dest_wb = openpyxl.load_workbook(dest_file)
      dest_ws = dest_wb.create_sheet(row['sheet_name'])
      
      # Write the code, title, subtitle1, subtitle2, source
      dest_ws.cell(row=1, column=1).value = row['code']
      dest_ws.cell(row=1, column=1).font = Font(name='Arial', size=8)
      dest_ws.cell(row=2, column=1).value = convert_footnote(row["ID"], row['title'])
      dest_ws.cell(row=2, column=1).font = Font(name='Arial', size=8)
      if(row['subtitle1'] != "None" and row['subtitle2'] != "None"):
        dest_ws.cell(row=3, column=1).value = convert_footnote(row["ID"], row['subtitle1'])
        dest_ws.cell(row=3, column=1).font = Font(name='Arial', size=8)
        dest_ws.cell(row=4, column=1).value = convert_footnote(row["ID"], row['subtitle2'])
        dest_ws.cell(row=4, column=1).font = Font(name='Arial', size=8)
        dest_ws.cell(row=6, column=1).value = "Quelle: " + row['source']
        dest_ws.cell(row=6, column=1).font = Font(name='Arial', size=8)
        title_toc = row['title_wfn'] + ", " + row['subtitle1_wfn'] + ", " + row['subtitle2_wfn']
        # define the row, where the content starts
        row_start = 8
      if(row['subtitle1'] != "None" and row['subtitle2'] == "None"):
        dest_ws.cell(row=3, column=1).value = row['subtitle1']
        dest_ws.cell(row=3, column=1).font = Font(name='Arial', size=8)
        dest_ws.cell(row=5, column=1).value = "Quelle: " + row['source']
        dest_ws.cell(row=5, column=1).font = Font(name='Arial', size=8)
        title_toc = row['title_wfn'] + ", " + row['subtitle1_wfn']
        # define the row, where the content starts
        row_start = 7
      if(row['subtitle1'] == "None" and row['subtitle2'] != "None"):
        dest_ws.cell(row=3, column=1).value = row['subtitle2']
        dest_ws.cell(row=3, column=1).font = Font(name='Arial', size=8)
        dest_ws.cell(row=5, column=1).value = "Quelle: " + row['source']
        dest_ws.cell(row=5, column=1).font = Font(name='Arial', size=8)
        title_toc = row['title_wfn'] + ", " + row['subtitle2_wfn']
        # define the row, where the content starts
        row_start = 7
      
      # Read the data from the source worksheet and write it to the destination worksheet
      # Copy all the format of the source worksheet "Internet"
      last_row = read_write_data(source_ws, dest_ws, row_start, row['ID'])

      # Get a list of all used footnotes for this worksheet
      list_wsfn = prepare_footnotes(row['ID'])
      row_start = last_row+3
      last_row = write_footnotes(dest_ws, list_wsfn, row_start)
      
      # Set the uniform row height to 12.75 for the common worksheet
      maxrow = dest_ws.max_row
      for r in range (1, maxrow + 100):
        dest_ws.row_dimensions[r].height = row_height

      # Write name and title in the table of contents
      content = dest_wb["Inhalt"]
      content.cell(row=count_toc, column=1).value = row['sheet_name']
      content.cell(row=count_toc, column=2).value = title_toc
      count_toc = count_toc + 1
      
      # saving the destination xlsx file
      dest_wb.save(dest_file)

# Function read_write_data
def read_write_data(source_ws, dest_ws, row_start, sheet_id):
  # set specific column width and hidden property
  # we cannot copy the entire column_dimensions attribute so we copy selected attributes
  for key, value in source_ws.column_dimensions.items():
    # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
    dest_ws.column_dimensions[key].min = copy(source_ws.column_dimensions[key].min)   
    dest_ws.column_dimensions[key].max = copy(source_ws.column_dimensions[key].max)
    # set width for every column
    dest_ws.column_dimensions[key].width = copy(source_ws.column_dimensions[key].width) 
    dest_ws.column_dimensions[key].hidden = copy(source_ws.column_dimensions[key].hidden)

  # calculate total number of rows and 
  # columns in source excel file
  mr = source_ws.max_row
  mc = source_ws.max_column

  # copying the cell values from source 
  # excel file to destination xlsx file
  for i in range (1, mr + 1):
    for j in range (1, mc + 1):
      # reading cell value from source xlsx
      c = source_ws.cell(row = i, column = j)

      # writing the read value to destination xlsx file
      # uses the convert_footnote function to convert any footnote references within string values
      if isinstance(c.value, str):
        dest_ws.cell(row = i+row_start, column = j).value = convert_footnote(sheet_id, c.value)
      else:
        dest_ws.cell(row = i+row_start, column = j).value = c.value
      
      # set the color to black and copy font
      # can throw deprecated warning: DeprecationWarning: Call to deprecated function copy (Use copy(obj) or cell.obj = cell.obj + other) 
      c.font = c.font.copy(color='FF000000')
      dest_ws.cell(row = i+row_start, column = j).font = copy(c.font)
      
      # copy alignment
      dest_ws.cell(row = i+row_start, column = j).alignment = copy(c.alignment)
      
      # copy number_format
      dest_ws.cell(row = i+row_start, column = j).number_format = copy(c.number_format)
  return(i+row_start)

# Main progam
def main():
  global data_coll
  
  tolog("INFO", "Read the configuration")
  read_config()
  
  tolog("INFO", "Loop trough all collection directories")
  read_coll_dir()
  
  tolog("INFO", "Open all excel sheets, read metadata and footnotes and add to dataframes")
  read_all_md_fn()
  
  tolog("INFO", "Loop over the collection and generating the tabsam")
  create_tabsam()
  

# Execute main of PyTabsam
if __name__ == '__main__':
  main()



